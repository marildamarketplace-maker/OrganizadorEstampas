from __future__ import annotations

from dataclasses import dataclass, asdict
from datetime import date, datetime
from pathlib import Path
from typing import Callable
import csv
import json
import re
import shutil
import tempfile
import time

from openpyxl import load_workbook, Workbook

from .config import COLUMN_ALIASES
from .indexer import image_key


@dataclass
class ProcessingItem:
    linha: int
    pedido: str
    data: str
    cliente: str
    base: str
    estampa: str
    variante: str
    arquivo_procurado: str
    status: str
    origem: str = ""
    destino: str = ""
    observacao: str = ""


@dataclass
class ProcessingSummary:
    total_linhas: int
    copiados: int
    nao_encontrados: int
    duplicados: int
    ignorados: int
    pedidos_criados: int
    elapsed_seconds: float
    report_xlsx: str
    report_csv: str


def normalize_header(value) -> str:
    text = "" if value is None else str(value)
    text = text.strip().casefold()
    text = re.sub(r"\s+", " ", text)
    return text


def clean_cell(value) -> str:
    if value is None:
        return ""
    # Evita transformar IDs numéricos inteiros em "123.0".
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_identifier(value) -> str:
    """Limpa um identificador e garante sua representação em maiúsculas."""
    return clean_cell(value).upper()


def image_search_names(estampa: str, variante: str) -> list[str]:
    """Retorna os nomes aceitos para uma estampa e sua variante."""
    names = [f"{estampa}-{variante}".strip("-")]
    # A variante A também pode ser armazenada sem o sufixo "-A".
    if variante == "A":
        names.append(estampa)
    return names


def exclusive_image_matches(
    index: dict[str, list[str]], searched_names: list[str]
) -> list[str]:
    """Busca arquivos exclusivos cujo código está no nome, não na pasta pai."""
    normalized_names = [name.casefold() for name in searched_names]
    matches: list[str] = []
    seen: set[str] = set()
    for key, paths in index.items():
        indexed_name = key.split("\0", maxsplit=1)[-1].casefold()
        if not any(
            indexed_name == name
            or indexed_name.startswith(name + " ")
            or indexed_name.startswith(name + ".")
            for name in normalized_names
        ):
            continue
        for path in paths:
            normalized_path = str(path).casefold()
            if normalized_path not in seen:
                seen.add(normalized_path)
                matches.append(str(path))
    return matches


def clean_order_date(value) -> tuple[str, str]:
    """Retorna a data para o relatório e uma versão segura para a pasta."""
    if isinstance(value, (datetime, date)):
        display = value.strftime("%d/%m/%Y")
        return display, value.strftime("%d-%m-%Y")

    display = clean_cell(value)
    if not display:
        return "", ""

    for date_format in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(display, date_format)
            return parsed.strftime("%d/%m/%Y"), parsed.strftime("%d-%m-%Y")
        except ValueError:
            pass

    return display, safe_folder_name(display)


def safe_folder_name(value: str) -> str:
    value = value.strip()
    value = re.sub(r'[<>:"/\\|?*]', "_", value)
    value = value.rstrip(". ")
    return value or "SEM_NOME"


def identifier_name_folder(identifier, name) -> str:
    """Monta pastas como CODIGO-NOME-DA-ENTIDADE."""
    identifier_part = clean_identifier(identifier)
    name_part = clean_identifier(name)
    name_part = re.sub(r"\s+", "-", name_part.strip())
    return safe_folder_name(f"{identifier_part}-{name_part}")


def discover_columns(headers: list) -> dict[str, int]:
    normalized = [normalize_header(h) for h in headers]
    discovered = {}

    for canonical, aliases in COLUMN_ALIASES.items():
        alias_set = {normalize_header(x) for x in aliases}
        for idx, header in enumerate(normalized):
            if header in alias_set:
                discovered[canonical] = idx
                break

    required = ["pedido", "data", "cliente", "base", "estampa", "variante"]
    missing = [name for name in required if name not in discovered]
    if missing:
        readable = {
            "pedido": "ID do Pedido",
            "data": "Data do Pedido",
            "cliente": "ID do Cliente",
            "base": "BASE",
            "estampa": "ID da Estampa",
            "variante": "Variante",
        }
        names = ", ".join(readable[x] for x in missing)
        raise ValueError(
            f"Não encontrei as colunas obrigatórias: {names}. "
            "Use a planilha modelo incluída no projeto."
        )

    return discovered


def process_excel(
    excel_path: Path,
    output_dir: Path,
    index: dict[str, list[str]],
    progress_callback: Callable[[int, int, str], None] | None = None,
) -> tuple[list[ProcessingItem], ProcessingSummary]:
    if not excel_path.exists():
        raise ValueError("O arquivo Excel selecionado não existe.")
    if excel_path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise ValueError("Use um arquivo Excel no formato .xlsx ou .xlsm.")
    if not output_dir.exists():
        output_dir.mkdir(parents=True, exist_ok=True)

    started = time.time()
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active

    rows = sheet.iter_rows(values_only=True)
    try:
        headers = list(next(rows))
    except StopIteration:
        raise ValueError("A planilha está vazia.")

    columns = discover_columns(headers)
    raw_rows = list(rows)
    workbook.close()
    total = len(raw_rows)
    results: list[ProcessingItem] = []
    pedidos = set()
    report_order_folders: set[Path] = set()

    copied = missing = duplicates = ignored = 0

    for position, row in enumerate(raw_rows, start=2):
        pedido = clean_identifier(row[columns["pedido"]] if columns["pedido"] < len(row) else "")
        estampa = clean_identifier(row[columns["estampa"]] if columns["estampa"] < len(row) else "")
        variante = clean_identifier(
            row[columns["variante"]] if columns["variante"] < len(row) else ""
        ) or "A"
        data, data_folder = clean_order_date(
            row[columns["data"]] if columns["data"] < len(row) else ""
        )
        cliente = clean_identifier(row[columns["cliente"]] if "cliente" in columns and columns["cliente"] < len(row) else "")
        base = clean_identifier(row[columns["base"]] if columns["base"] < len(row) else "")

        searched_names = image_search_names(estampa, variante)
        searched_name = " ou ".join(searched_names)
        item = ProcessingItem(
            linha=position,
            pedido=pedido,
            data=data,
            cliente=cliente,
            base=base,
            estampa=estampa,
            variante=variante,
            arquivo_procurado=searched_name,
            status="",
        )

        if not pedido or not data or not cliente or not base or not estampa:
            item.status = "IGNORADO"
            item.observacao = "Pedido, data, cliente, base ou estampa em branco."
            ignored += 1
            results.append(item)
            continue

        report_order_folders.add(
            output_dir
            / safe_folder_name(cliente)
            / data_folder
            / safe_folder_name(pedido)
        )

        matches = [
            match
            for candidate in searched_names
            for match in index.get(image_key(estampa, candidate), [])
        ]
        exclusive_matches = False
        if not matches:
            matches = exclusive_image_matches(index, searched_names)
            exclusive_matches = bool(matches)

        if not matches:
            item.status = "NÃO ENCONTRADO"
            item.observacao = (
                f"Nenhuma imagem para '{estampa}/{searched_name}'."
            )
            missing += 1
        elif len(matches) > 1 and not exclusive_matches:
            item.status = "DUPLICADO"
            item.origem = " | ".join(matches)
            item.observacao = "Há mais de um arquivo com o mesmo nome. Nada foi copiado."
            duplicates += 1
        else:
            order_folder = (
                output_dir
                / safe_folder_name(cliente)
                / data_folder
                / safe_folder_name(pedido)
                / safe_folder_name(base)
            )
            order_folder.mkdir(parents=True, exist_ok=True)
            destinations: list[str] = []
            copied_for_item = 0
            existing_for_item = 0
            for match in matches:
                source = Path(match)
                destination = order_folder / source.name.upper()
                destinations.append(str(destination))
                if destination.exists():
                    existing_for_item += 1
                else:
                    shutil.copy2(source, destination)
                    copied_for_item += 1

            item.origem = " | ".join(matches)
            item.destino = " | ".join(destinations)
            copied += copied_for_item
            ignored += existing_for_item
            if copied_for_item:
                item.status = "COPIADO"
                item.observacao = f"{copied_for_item} arquivo(s) copiado(s)."
                if existing_for_item:
                    item.observacao += f" {existing_for_item} já existente(s)."
                pedidos.add((cliente, data_folder, pedido))
            else:
                item.status = "JÁ EXISTE"
                item.observacao = "Todos os arquivos já existem no pedido."

        results.append(item)

        if progress_callback:
            current = position - 1
            progress_callback(current, total, f"Processando linha {current} de {total}")

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    report_dir = (
        next(iter(report_order_folders))
        if len(report_order_folders) == 1
        else output_dir
    )
    report_dir.mkdir(parents=True, exist_ok=True)
    report_xlsx = report_dir / f"relatorio_processamento_{timestamp}.xlsx"
    report_csv = report_dir / f"relatorio_processamento_{timestamp}.csv"
    write_reports(results, report_xlsx, report_csv)

    summary = ProcessingSummary(
        total_linhas=total,
        copiados=copied,
        nao_encontrados=missing,
        duplicados=duplicates,
        ignorados=ignored,
        pedidos_criados=len(pedidos),
        elapsed_seconds=time.time() - started,
        report_xlsx=str(report_xlsx),
        report_csv=str(report_csv),
    )
    return results, summary


def process_csv_text(
    csv_text: str,
    output_dir: Path,
    index: dict[str, list[str]],
    progress_callback: Callable[[int, int, str], None] | None = None,
) -> tuple[list[ProcessingItem], ProcessingSummary]:
    """Processa texto CSV, com ou sem cabeçalho, usando a mesma regra do Excel."""
    text = csv_text.strip().lstrip("\ufeff")
    if not text:
        raise ValueError("Cole os pedidos no campo de texto CSV.")

    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if "\t" in text else ";" if ";" in text else ","

    parsed_rows = [
        [cell.strip() for cell in row]
        for row in csv.reader(text.splitlines(), delimiter=delimiter)
        if any(cell.strip() for cell in row)
    ]
    if not parsed_rows:
        raise ValueError("O texto CSV não contém pedidos.")

    expected_headers = [
        "ID do Pedido",
        "Data do Pedido",
        "ID do Cliente",
        "BASE",
        "ID da Estampa",
        "Variante",
    ]

    try:
        discover_columns(parsed_rows[0])
        rows_to_write = parsed_rows
    except ValueError:
        if len(parsed_rows[0]) < len(expected_headers):
            raise ValueError(
                "O texto CSV deve ter 6 colunas nesta ordem: "
                + ", ".join(expected_headers)
            )
        rows_to_write = [expected_headers, *parsed_rows]

    workbook = Workbook()
    sheet = workbook.active
    for row in rows_to_write:
        sheet.append(row)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temporary:
        temporary_path = Path(temporary.name)
    try:
        workbook.save(temporary_path)
        workbook.close()
        return process_excel(
            temporary_path,
            output_dir,
            index,
            progress_callback=progress_callback,
        )
    finally:
        temporary_path.unlink(missing_ok=True)


def process_order_payload(
    payload: dict,
    output_dir: Path,
    index: dict[str, list[str]],
    progress_callback: Callable[[int, int, str], None] | None = None,
) -> tuple[list[ProcessingItem], ProcessingSummary]:
    """Processa o JSON extraído de um PDF de pedido, sem depender da interface."""
    if not isinstance(payload, dict):
        raise ValueError("O conteúdo do pedido deve ser um objeto JSON.")

    required = ("pedido", "data", "clienteCodigo", "clienteNome", "produtos")
    missing = [field for field in required if not payload.get(field)]
    if missing:
        raise ValueError(
            "Campos obrigatórios ausentes no JSON: " + ", ".join(missing)
        )

    products = payload["produtos"]
    if not isinstance(products, list) or not products:
        raise ValueError("O campo 'produtos' deve conter pelo menos um produto.")

    rows = []
    base_folders: set[str] = set()
    product_fields = ("tecidoCodigo", "tecidoNome", "estampa")
    for position, product in enumerate(products, start=1):
        if not isinstance(product, dict):
            raise ValueError(f"O produto {position} deve ser um objeto JSON.")
        product_missing = [field for field in product_fields if not product.get(field)]
        if product_missing:
            raise ValueError(
                f"Campos ausentes no produto {position}: "
                + ", ".join(product_missing)
            )
        product_folder = identifier_name_folder(
            product["tecidoCodigo"], product["tecidoNome"]
        )
        customer_folder = identifier_name_folder(
            payload["clienteCodigo"], payload["clienteNome"]
        )
        rows.append([
            payload["pedido"],
            payload["data"],
            customer_folder,
            product_folder,
            product["estampa"],
            clean_identifier(product.get("variante", "")) or "A",
        ])
        base_folders.add(product_folder)

    pedido = clean_identifier(payload["pedido"])
    cliente_folder = identifier_name_folder(
        payload["clienteCodigo"], payload["clienteNome"]
    )
    _, data_folder = clean_order_date(payload["data"])
    order_folder = (
        output_dir
        / cliente_folder
        / data_folder
        / safe_folder_name(pedido)
    )
    # A pasta do pedido deve existir mesmo quando alguma imagem não for localizada.
    for base in base_folders:
        (order_folder / safe_folder_name(base)).mkdir(parents=True, exist_ok=True)

    csv_text = "\n".join(
        ";".join(clean_cell(value) for value in row) for row in rows
    )
    results, summary = process_csv_text(
        csv_text,
        output_dir,
        index,
        progress_callback=progress_callback,
    )
    summary.pedidos_criados = 1
    return results, summary


def process_order_json(
    json_text: str,
    output_dir: Path,
    index: dict[str, list[str]],
    progress_callback: Callable[[int, int, str], None] | None = None,
) -> tuple[list[ProcessingItem], ProcessingSummary]:
    """Converte texto JSON e cria a estrutura do pedido."""
    try:
        payload = json.loads(json_text.lstrip("\ufeff"))
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"JSON inválido na linha {exc.lineno}, coluna {exc.colno}: {exc.msg}"
        ) from exc
    return process_order_payload(payload, output_dir, index, progress_callback)


def write_reports(results: list[ProcessingItem], xlsx_path: Path, csv_path: Path):
    headers = list(asdict(results[0]).keys()) if results else [
        "linha", "pedido", "data", "cliente", "base", "estampa", "variante",
        "arquivo_procurado", "status", "origem", "destino", "observacao"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"
    ws.append(headers)
    for item in results:
        ws.append([asdict(item)[h] for h in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 10, "B": 22, "C": 16, "D": 20, "E": 18, "F": 16,
        "G": 14, "H": 24, "I": 18, "J": 55, "K": 55, "L": 45
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    wb.save(xlsx_path)

    with csv_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=headers, delimiter=";")
        writer.writeheader()
        for item in results:
            writer.writerow(asdict(item))
